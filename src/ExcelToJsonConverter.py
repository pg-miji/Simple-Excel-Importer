import openpyxl
import json
import os

inputPath = './excel/'
outputPath = './json/'

files = os.listdir(inputPath)
for file in files:

    fileName = file.replace('.xlsx', '')
    if not os.path.exists(outputPath + fileName):
        os.mkdir(outputPath + fileName)

    excel = openpyxl.load_workbook(inputPath)
    sheetNames = excel.sheetnames
    if 'unnecessary file' in sheetNames:
        sheetNames.remove('unnecessary file')

    for sheetName in sheetNames:
        sheet = excel[sheetName]
        print('sheet: '+sheetName)

        for i in range(11, sheet.max_row):

            name = sheet.cell(row=i, column=2).value
            comment = sheet.cell(row=i, column=8).value
            column = {
                "name": name,
                "comment": str(comment)
            }

            with open(outputPath+'/'+sheetName+'.json', 'a', encoding="utf-8") as f:
                if i == 11:
                    f.write('[')
                elif name == '' or name is None:
                    f.write(']')
                    break
                else:
                    f.write(',')

                f.write(json.dumps(column, ensure_ascii=False, indent=4))
